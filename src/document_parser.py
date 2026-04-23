"""document_parser.py — Multi-format document parsing for ingestion.

Supports: PDF, Excel (.xlsx), CSV, Word (.docx), TXT, Markdown
Extracts text with structure awareness for financial docs and meeting transcripts.
"""
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from typing import Optional


def parse_document(file_path: str) -> dict:
    """Parse any supported document format and return structured text.

    Returns:
        dict with:
            - text: full extracted text
            - chunks: list of text chunks (for embedding)
            - metadata: source, page_count, etc.
            - format: detected format
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".pdf":
        return _parse_pdf(file_path)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(file_path)
    elif ext == ".csv":
        return _parse_csv(file_path)
    elif ext in (".docx", ".doc"):
        return _parse_docx(file_path)
    elif ext in (".txt", ".md", ".markdown"):
        return _parse_text(file_path)
    else:
        return {"text": "", "chunks": [], "metadata": {}, "format": "unknown", "error": f"Unsupported format: {ext}"}


def _parse_pdf(path: str) -> dict:
    """Extract text from PDF with page awareness."""
    import pdfplumber

    pages = []
    full_text = []

    try:
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                if text.strip():
                    pages.append({"page": i + 1, "text": text})
                    full_text.append(f"[Page {i+1}]\n{text}")

        # Chunk pages for better context (2 pages per chunk)
        chunks = []
        for i in range(0, len(pages), 2):
            chunk_pages = pages[i:i+2]
            chunk_text = "\n\n".join(p["text"] for p in chunk_pages)
            if chunk_text.strip():
                chunks.append({
                    "text": chunk_text,
                    "pages": [p["page"] for p in chunk_pages],
                    "source": f"{os.path.basename(path)}:p{chunk_pages[0]['page']}-{chunk_pages[-1]['page']}"
                })

        return {
            "text": "\n\n".join(full_text),
            "chunks": chunks,
            "metadata": {
                "page_count": len(pages),
                "source": os.path.basename(path),
            },
            "format": "pdf",
        }
    except Exception as e:
        return {"text": "", "chunks": [], "metadata": {}, "format": "pdf", "error": str(e)}


def _parse_excel(path: str) -> dict:
    """Extract text from Excel with sheet awareness."""
    import openpyxl

    sheets = []
    full_text = []

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Convert row to string, filter empty
                row_str = " | ".join([str(c) if c is not None else "" for c in row])
                if row_str.strip():
                    rows.append(row_str)

            sheet_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
            sheets.append({"sheet": sheet_name, "row_count": len(rows), "text": sheet_text})
            full_text.append(sheet_text)

        # Chunk by sheet
        chunks = [{"text": s["text"], "sheet": s["sheet"], "source": f"{os.path.basename(path)}/{s['sheet']}"} for s in sheets]

        return {
            "text": "\n\n".join(full_text),
            "chunks": chunks,
            "metadata": {"sheet_count": len(sheets), "source": os.path.basename(path)},
            "format": "excel",
        }
    except Exception as e:
        return {"text": "", "chunks": [], "metadata": {}, "format": "excel", "error": str(e)}


def _parse_csv(path: str) -> dict:
    """Parse CSV as structured data."""
    import pandas as pd

    try:
        df = pd.read_csv(path, dtype=str, keep_default_na=False)
        total_rows = len(df)
        # Paginate large CSVs: emit chunks of 100 rows each to avoid truncation
        chunk_size = 100
        lines = [f"[CSV: {os.path.basename(path)}]", f"Columns: {', '.join(df.columns.tolist())}", ""]
        truncated_note = ""
        if total_rows > chunk_size:
            truncated_note = f" (showing {total_rows} rows in chunks)"
        lines.append(f"Total rows: {total_rows}{truncated_note}")
        lines.append("")

        for start in range(0, min(total_rows, 1000), chunk_size):
            end = min(start + chunk_size, total_rows)
            chunk_df = df.iloc[start:end]
            lines.append(f"--- Rows {start+1}-{end} ---")
            lines.append(chunk_df.to_string(max_rows=100, max_colwidth=50))
            lines.append("")

        text = "\n".join(lines)
        chunks = [{"text": text, "source": os.path.basename(path)}]

        return {
            "text": text,
            "chunks": chunks,
            "metadata": {"row_count": len(df), "columns": df.columns.tolist(), "source": os.path.basename(path)},
            "format": "csv",
        }
    except Exception as e:
        return {"text": "", "chunks": [], "metadata": {}, "format": "csv", "error": str(e)}


def _parse_docx(path: str) -> dict:
    """Extract text from Word documents."""
    from docx import Document

    try:
        doc = Document(path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

        full_text = "\n".join(paragraphs)
        # Chunk by 5 paragraphs
        chunks = []
        for i in range(0, len(paragraphs), 5):
            chunk_paras = paragraphs[i:i+5]
            chunk_text = "\n".join(chunk_paras)
            if chunk_text.strip():
                chunks.append({
                    "text": chunk_text,
                    "source": f"{os.path.basename(path)}:p{i//5+1}"
                })

        return {
            "text": full_text,
            "chunks": chunks,
            "metadata": {"paragraph_count": len(paragraphs), "source": os.path.basename(path)},
            "format": "docx",
        }
    except Exception as e:
        return {"text": "", "chunks": [], "metadata": {}, "format": "docx", "error": str(e)}


def _parse_text(path: str) -> dict:
    """Parse plain text and markdown files."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()

        # Simple chunking by paragraph (split on double newline)
        paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
        chunks = []
        for i in range(0, len(paragraphs), 5):
            chunk = paragraphs[i:i+5]
            chunk_text = "\n\n".join(chunk)
            if chunk_text.strip():
                chunks.append({
                    "text": chunk_text,
                    "source": f"{os.path.basename(path)}:chunk{i//5+1}"
                })

        return {
            "text": text,
            "chunks": chunks,
            "metadata": {"source": os.path.basename(path), "char_count": len(text)},
            "format": "text",
        }
    except Exception as e:
        return {"text": "", "chunks": [], "metadata": {}, "format": "text", "error": str(e)}


# ---------------------------------------------------------------------------
# Chunking utilities
# ---------------------------------------------------------------------------

def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> list[dict]:
    """Split long text into overlapping chunks for embedding.

    Args:
        text: Text to chunk
        chunk_size: Target characters per chunk
        overlap: Character overlap between chunks

    Returns:
        list of {"text": str, "source": str} dicts
    """
    if len(text) <= chunk_size:
        return [{"text": text.strip(), "source": "inline"}]

    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        chunks.append({"text": chunk.strip(), "source": f"chars_{start}_{end}"})
        start += chunk_size - overlap

    return [c for c in chunks if c["text"]]


if __name__ == "__main__":
    import tempfile

    # Test with a simple text file
    print("Testing document parser...")

    # Create a test file
    test_content = """
    Financial Summary Q1 2024

    Revenue: $1.2M
    Expenses: $800K
    Net Income: $400K

    Key highlights:
    - Product sales increased 25% year over year
    - Operating costs remained flat
    - New market expansion in APAC region
    """

    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as f:
        f.write(test_content)
        tmp = f.name

    result = parse_document(tmp)
    print(f"Format: {result['format']}")
    print(f"Text length: {len(result['text'])}")
    print(f"Chunks: {len(result['chunks'])}")
    print(f"First chunk: {result['chunks'][0]['text'][:100] if result['chunks'] else 'none'}...")

    os.unlink(tmp)
    print("Parser working!")